# -*- coding: utf-8 -*-
from __future__ import unicode_literals

from django.shortcuts import render, redirect, HttpResponse
from .models import *
import mimetypes
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook
from io import BytesIO
from .forms import *
# Create your views here.

csrf_exempt
@login_required(login_url='/loginAcc')
def homepage(request):
	form = UploadForm()
	user = request.user
	totaData = DataModel.objects.all().count()
	userAccount = MyUser.objects.get(user = user)
	lastAddOn = DataModel.objects.latest('id')
	notifications = NotificationMaster.objects.all().order_by('-id')
	addsearches = addAsearch.objects.all().order_by('-id')
	loadPort = list(set(DataModel.objects.values_list('loadPortDefinition', flat=True).distinct()))
	dischargePort = list(set(DataModel.objects.values_list('dischargePortDefinition', flat=True).distinct()))

	return render(request,'index.html',{'dischargePort':dischargePort, 'loadPort':loadPort, 'addsearches':addsearches, 'lastAddOn':lastAddOn, 'totalData':totaData, 'form':form,'userAccount':userAccount,'notifications':notifications})

@csrf_exempt
def loginAcc(request):

	if request.method == 'POST':
		form = request.POST
		username = form['username']
		password = form['password']
		
		user = authenticate(username = username,password = password)
		
		if user is not None:
			# if userAccount.is_activeYesNo == True:
					
			login(request,user)
			return redirect('/home')
			
		else:
			message = 'Enter correct username and password!!!'
			return render(request,'login.html',{'message':message})
		# except:
		# 	message = 'Your account not found!!!'
		# 	return render(request,'login.html',{'message':message})

	else:
		return render(request,'login.html')

@csrf_exempt
@login_required(login_url='/loginAcc')
def user_logout(request):
	user = request.user
	logout(request)
	return redirect('/loginAcc')


import openpyxl

csrf_exempt
@login_required(login_url='/loginAcc')
def fileUpload(request):
	user = request.user
	userAccount = MyUser.objects.get(user = user)
	if "GET" == request.method:
		return render(request, 'index.html', {})
	else:
		excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

		wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
		worksheet = wb["Sheet1"]
		
		count = 0
        # iterating over the rows and
        # getting value from each cell in row
		for row in worksheet.iter_rows():
			count += 1
			row_data = list()
			for cell in row:
				row_data.append(str(cell.value))

			# print(float(row_data[1]))

			#upload details to database
			datamodelobject = DataModel.objects.create(
				companyId = row_data[0],
				operationFormID = row_data[1],
				voyageId = row_data[2],
				BLdate = row_data[3],
				vesselId = row_data[4],
				VesselName = row_data[5],
				buyerId = row_data[6],
				buyerName = row_data[7],
				sellerId = row_data[8],
				sellerName = row_data[9],
				producerId = row_data[10],
				producerName = row_data[11],
				loadPortId = row_data[12],
				loadPortDefinition = row_data[13],
				loadCountryCode = row_data[14],
				loadCountry = row_data[15],
				dischargePortId = row_data[16],
				dischargePortDefinition = row_data[17],
				dischargeCountryCode = row_data[18],
				dischargeCountry = row_data[19],
				productId = row_data[20],
				productDefinition = row_data[21],
				operationLine = row_data[22],
				packingType = row_data[23],
				packingTypeDefinition = row_data[24],
				totalQuantity = row_data[25],
				ETALoadPort = row_data[26],
				lay = row_data[27],
				can = row_data[28],
				purchaseIncoterm = row_data[29],
				purchaseIncotermDefinition = row_data[30],
				salesIncoterm = row_data[31],
				salesIncotermDefinition = row_data[32],
				shipResp = row_data[33],
				shipRespDesc = row_data[34],
				tradeOperationsResponsible = row_data[35],
				tradeOperationsResponsibleDefinition = row_data[36],
				addUser = row_data[37],
				addDate = row_data[38],
				purchaseContractNo = row_data[39],
				purchaseContractLine = row_data[40],
				salesContractNo = row_data[41],
				salesContractLine = row_data[42],
				approvalId = row_data[43],
				approvalSubjectStatus = row_data[44],
				freightPMT = row_data[45],
				freightCurrency = row_data[46],
				loadingRate = row_data[47],
				dischargeRate = row_data[48],
				netBLQuantity = row_data[49],
				purchasePrice = row_data[50],
				salesPrice = row_data[51],
				netCommision = row_data[52],
				sample = row_data[53]

				)


		# 	excel_data.append(row_data)
		# print(excel_data)

		notificationIs = NotificationMaster.objects.create(
			notification_is = 'Hi '+ str(user.first_name) +', You have added ' + str(count) +' new data to database.',
			notification_by = userAccount
			)
		return redirect('/home')

csrf_exempt
@login_required(login_url='/loginAcc')
def addsearch(request):
	if request.method == 'POST':
		form = request.POST

		fieldName = form['fieldName']
		datatype = form['datatype']

		addsearchmodel = addAsearch.objects.create(
			fieldIs = fieldName,
			fieldType = datatype
			)

		return redirect('/home')

	else:
		return redirect('/home')

csrf_exempt
@login_required(login_url='/loginAcc')
def searchByField(request):
	loadPort = list(set(DataModel.objects.values_list('loadPortDefinition', flat=True).distinct()))
	dischargePort = list(set(DataModel.objects.values_list('dischargePortDefinition', flat=True).distinct()))

	if request.method == 'POST':
		form = request.POST

		loadPorts = form['loadPort']
		dischargePorts = form['dischargePort']

		if loadPorts is not None and dischargePorts is None:
			searched = DataModel.objects.filter(loadPortDefinition__icontains = loadPorts)
		elif loadPorts is None and dischargePorts is not None:
			searched = DataModel.objects.filter(dischargePortDefinition__icontains = dischargePorts)
		elif loadPorts is not None and dischargePorts is not None:
			searched = DataModel.objects.filter(loadPortDefinition__icontains = loadPorts, dischargePortDefinition__icontains = dischargePorts,)

		return render(request,'table-basic.html',{'searched':searched, 'dischargePort':dischargePort, 'loadPort':loadPort})
	else:
		return HttpResponse('<h1>Error 404 </h1>')

csrf_exempt
@login_required(login_url='/loginAcc')
def allData(request):
	searched = DataModel.objects.all()
	return render(request,'table-basic.html',{'searched':searched})
